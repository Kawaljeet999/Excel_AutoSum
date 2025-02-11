import openpyxl

# Create a new workbook and select the active sheet
wb = openpyxl.Workbook()
ws = wb.active

# Assign the value 12 to every cell in column 1 (rows 1 to 10)
for row in range(1, 11):
    ws.cell(row=row, column=1, value=12)

# Assign different values to column 2 (rows 1 to 10)
values = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10]

for row, value in enumerate(values, start=1): # start index from row 1 
    ws.cell(row=row, column=2, value=value)
    
# Iterate through the rows where the data is present
for row in range(1, 11):  # Adjust the range according to the number of rows you have
    # Get the values from columns A and B
    value_a = ws.cell(row=row, column=1).value  # Column 1 is A
    value_b = ws.cell(row=row, column=2).value  # Column 2 is B
  
    # Calculate the sum
    sum_value = value_a + value_b
    
    # Store the result in column C (Column 3)
    ws.cell(row=row, column=3, value=sum_value)

# Alither
# for i in range(1,11):
#     ws[f'C{i}'] = f'=A{i}+B{i}'

# Save the workbook
wb.save('Column_ADD.xlsx')
